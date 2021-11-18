#!/usr/bin/python3.7
import optparse
import datetime
import pdb
from requests.auth import HTTPBasicAuth
import requests
import smtplib
import sys
import re
import json
from datetime import date
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from os.path import basename

from openpyxl import Workbook


ALL_PROJECTS = ["GPE","FSE","FOO"]


def get_resource(story, user, api_token):  
    data = get_resource_all(story, user, api_token)
   
    tpr = ResourceData()
    

    if "aggregatetimeoriginalestimate" in data['fields']:
        aggr = int(data['fields']['aggregatetimeoriginalestimate']) if data['fields']['aggregatetimeoriginalestimate'] else 0 
        tpr.original_estimation = aggr/60/60
    if "timeSpentSeconds" in data['fields']['timetracking']:
        tpr.hours_spent = int(data['fields']['timetracking']['timeSpentSeconds'])/60/60
    if "updated" in data['fields']:
        tpr.last_update = data['fields']['updated']
    if "name" in data['fields']['status']:
        tpr.status = data['fields']['status']['name']
    if "customfield_10405" in data['fields'] and data['fields']['customfield_10405'] is not None:
        tpr.story_point = int(data['fields']['customfield_10405'])
    if "labels" in data['fields']:
        tpr.labels = data['fields']['labels']  
    if "subtasks" in data['fields']:
        for sub in data['fields']['subtasks']:
            tpr.subtasks+=sub['key']+" "  
    return tpr



class ResourceData:
    def __init__(self):
        self.hours_spent = 0
        self.last_update = ""
        self.status = ""
        self.story_point = 0
        self.labels = []
        self.original_estimation = 0
        self.subtasks= ""
    
    def serialize(self):
        return{
            'hours_spent' : self.hours_spent,
            'last_update' : self.last_update,
            'status' : self.status,
            'estimate' : self.story_point
        }
   
class Issue:
    def __init__(self):
        self.id = ""
        self.summary = ""
        self.resource = ResourceData()
    def serialize(self):
        return {
            'id': self.id, 
            'summary': self.summary,
            'resource': self.resource.serialize()
            
        }

def get_issues(project_id, user, api_token):
    url = 'https://#JIRAURL#/rest/api/3/issue/picker'
        
    auth = HTTPBasicAuth(user, api_token)
    
    headers = {
    "Accept": "application/json",
    "Content-Type": "application/json"
    }
    query = {
        'currentProjectId': '{}'.format(project_id),
        'currentJQL' :  'order by status DESC',
        'showSubTasks' : True
        }
    
    response = requests.request(
                        "GET",
                        url,
                        headers=headers,
                        params=query,
                        auth=auth)

    
    return response.json()

def search_issues(project_id, user, api_token):
    url = "https://#JIRAURL#/rest/api/3/search"

    auth = HTTPBasicAuth(user, api_token)
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }


    query = {
        'jql': "project = {}".format(project_id),
        'maxResults': "5000"
    }


    response = requests.request(
                        "GET",
                        url,
                        headers=headers,
                        params=query,
                        auth=auth)

    return response.json()


def get_resource_all(story, user, api_token):
    r = requests.get('https://#JIRAURL#/rest/api/latest/issue/{}'.format(story), auth=HTTPBasicAuth(user, api_token))
    return r.json()



parser = optparse.OptionParser()

origin_est_lst = []
worklog_lst = []

parser.add_option('-d', '--data',
    action="store", dest="from_data",
    help="from data")



issues_list = []

options, args = parser.parse_args()

USER = "#USER#"
API_TOKEN = "#TOKEN#"


if options.from_data:
    from_date = options.from_data
else:
    from_date = ""


today = date.today()


FILE_NAME = "task_report.xls"
wb = Workbook()

status_list = {}
for proj_id in ALL_PROJECTS:
    data = search_issues(proj_id, USER, API_TOKEN)
    i = 0
    for issue in data['issues']:
        
        if from_date!="":
            created = datetime.datetime.strptime(issue['fields']['created'].split('T')[0], '%Y-%m-%d')
            startdate = datetime.datetime.strptime(from_date, '%Y-%m-%d')
            if created >= startdate:
                id = issue['key']
                status = issue['fields']['status']['name'].lower()
                
                actual_sheet = None
                actual_row = 1
                if status not in status_list:
                    actual_sheet = wb.create_sheet(title=status)
                    
                    #write header (only for the first time)
                    col = 1
                    actual_sheet.cell(column=col, row=1, value="id")
                    col = col+1
                    actual_sheet.cell(column=col, row=1, value="summary")
                    col = col+1
                    actual_sheet.cell(column=col, row=1, value="status")
                    col = col+1
                    actual_sheet.cell(column=col, row=1, value="created")
                    col = col+1
                    actual_sheet.cell(column=col, row=1, value="worklog")
                    col = col+1
                    actual_sheet.cell(column=col, row=1, value="original_estimation")
                    col = col+1
                    actual_sheet.cell(column=col, row=1, value="story_point")
                    col = col+1
                    actual_sheet.cell(column=col, row=1, value="strsubtasks")
                    col = col+1
                    actual_sheet.cell(column=col, row=1, value="assignee")

                    status_list[status] = [1,actual_sheet]
                else:
                    actual_row = status_list[status][0]
                    actual_sheet = status_list[status][1]
                    

                summary = issue['fields']['summary'][0:100].encode('utf-8')
                
                if "timetracking" in issue['fields'] and "timeSpentSeconds" in issue['fields']['timetracking']:
                    worklog = (int(issue['fields']['timetracking']['timeSpentSeconds']))/60/60
                else:
                    worklog = 0

                if "aggregatetimeoriginalestimate" in issue['fields']:
                    original_estimation = int(int(issue['fields']['aggregatetimeoriginalestimate']) if issue['fields']['aggregatetimeoriginalestimate'] else 0)/60/60
                else:   
                    original_estimation = 0

                if "customfield_10405" in issue['fields'] and issue['fields']['customfield_10405'] is not None:
                    story_point = int(issue['fields']['customfield_10405'])
                else:
                    story_point = 0

                try:
                    assignee = issue['fields']['assignee']['displayName']
                except:
                    assignee = "N/A"

                subtasks = []
                if "subtasks" in issue['fields']:
                    for sub in issue['fields']['subtasks']:
                        subtasks.append(sub['key'])
                if len(subtasks)>0:
                    strsubtasks = " ".join(subtasks)
                else:
                    strsubtasks = "N/A"

                actual_row = actual_row+1
                actual_sheet.cell(column=1, row=actual_row, value=id)
                actual_sheet.cell(column=2, row=actual_row, value=summary)
                actual_sheet.cell(column=3, row=actual_row, value=status)
                actual_sheet.cell(column=4, row=actual_row, value=created)
                actual_sheet.cell(column=5, row=actual_row, value=worklog)
                actual_sheet.cell(column=6, row=actual_row, value=original_estimation)
                actual_sheet.cell(column=7, row=actual_row, value=story_point)
                actual_sheet.cell(column=8, row=actual_row, value=strsubtasks)
                actual_sheet.cell(column=9, row=actual_row, value=assignee)
                status_list[status][0] = actual_row
wb.save(filename = FILE_NAME)


gmail_user = '#EMAIL#'
gmail_password = '#PASSWORD#'
to_mails = "#TOEMAIL#"
sent_from = gmail_user
to =[]
for email in to_mails.split(","):
    to.append(email.strip())


body=""
subject = "Estrazione avanzamento task"
body = "Estrazione avanzamento task"
    


msg = MIMEMultipart()
msg['From'] = gmail_user
msg['To'] = COMMASPACE.join(to_mails)
msg['Date'] = formatdate(localtime=True)
msg['Subject'] = subject

msg.attach(MIMEText(body))

with open(FILE_NAME, "rb") as fil:
    part = MIMEApplication(
        fil.read(),
        Name=basename(FILE_NAME)
    )
# After the file is closed
part['Content-Disposition'] = 'attachment; filename="%s"' % basename(FILE_NAME)
msg.attach(part)


server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
server.ehlo()
server.login(gmail_user, gmail_password)
server.sendmail(sent_from, to, msg.as_string())
server.close()

print('Email sent!')

            







    


    

